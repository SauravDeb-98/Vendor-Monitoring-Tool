"""
Data Ingestion Module
----------------------
Parses an uploaded Excel vendor list. Per spec, this module STRICTLY extracts
only two columns from the source file: Vendor Name and Vendor Website Link.
Any other columns present in the source workbook are ignored. Input is
cleaned/standardized before being handed to the scanner.

Cell reading goes through openpyxl directly (not pandas) for the data rows,
because pandas' read_excel returns cached calculated values for formula
cells and silently yields blank/NaN when a workbook has no cached value for
a formula (common in programmatically-generated files). This affects, for
example, =HYPERLINK("url", "label") formulas and real Excel hyperlink
objects attached to cells — both are resolved explicitly below.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass
from urllib.parse import urlparse

import openpyxl
import pandas as pd

# Column header aliases we will recognize in the source file, matched
# case-insensitively with whitespace stripped.
NAME_ALIASES = {"vendor name", "vendor", "name", "company", "company name", "supplier", "supplier name"}
URL_ALIASES = {
    "vendor website link", "vendor website", "website", "website link", "url",
    "web url", "domain", "site", "vendor url", "link",
}


@dataclass
class Vendor:
    name: str
    website: str
    domain: str


class IngestionError(ValueError):
    pass


def _normalize_header(h) -> str:
    return re.sub(r"\s+", " ", str(h).strip().lower())


def _find_column_index(header_row_values: list, aliases: set) -> int | None:
    normalized = [(_normalize_header(v) if v is not None else "") for v in header_row_values]
    for idx, norm in enumerate(normalized):
        if norm in aliases:
            return idx
    for idx, norm in enumerate(normalized):
        if norm and any(alias in norm or norm in alias for alias in aliases):
            return idx
    return None


_HYPERLINK_FORMULA_RE = re.compile(r'=HYPERLINK\(\s*"([^"]+)"', re.IGNORECASE)


def _extract_cell_url(cell) -> str | None:
    """
    Resolve a URL from a worksheet cell that may contain:
      - a plain text URL/domain
      - a real Excel hyperlink object (cell.hyperlink.target)
      - an =HYPERLINK("url", "label") formula (as literal text, since the
        workbook may have no cached calculated value)
    """
    if cell.hyperlink and cell.hyperlink.target:
        return str(cell.hyperlink.target).strip()
    value = cell.value
    if value is None:
        return None
    text = str(value).strip()
    formula_match = _HYPERLINK_FORMULA_RE.match(text)
    if formula_match:
        return formula_match.group(1).strip()
    return text


def clean_url(raw: str | None) -> str | None:
    """Public wrapper around the same URL-cleaning logic used during Excel
    ingestion, so other entry points (e.g. the single-vendor "Add a Vendor"
    flow in main.py) validate/normalize a website URL identically rather
    than duplicating this logic."""
    return _clean_url(raw)


def extract_domain(url: str) -> str:
    """Public wrapper around the same domain-extraction logic used during
    Excel ingestion; see clean_url's docstring for why this is exposed."""
    return _extract_domain(url)


def _clean_url(raw: str | None) -> str | None:
    if not raw or not isinstance(raw, str):
        return None
    raw = raw.strip()
    if not raw or raw.lower() in {"n/a", "na", "none", "-"}:
        return None
    if not re.match(r"^https?://", raw, re.IGNORECASE):
        raw = "https://" + raw
    parsed = urlparse(raw)
    if not parsed.netloc or "." not in parsed.netloc:
        return None
    # Strip path/query — we want the site root for scanning purposes
    clean = f"{parsed.scheme}://{parsed.netloc}".lower()
    return clean


def _extract_domain(url: str) -> str:
    netloc = urlparse(url).netloc.lower()
    return netloc[4:] if netloc.startswith("www.") else netloc


def _find_header_row(ws, max_scan_rows: int = 15) -> int | None:
    """
    Scan the first N rows of the sheet looking for the row that contains
    both a recognizable name-column header and a recognizable url-column
    header. Returns the 1-indexed openpyxl row number, or None if not found.
    This handles files with decorative title/banner rows above the real
    header row.
    """
    for row_idx in range(1, max_scan_rows + 1):
        row_values = [cell.value for cell in ws[row_idx]]
        if _find_column_index(row_values, NAME_ALIASES) is not None and \
           _find_column_index(row_values, URL_ALIASES) is not None:
            return row_idx
    return None


def parse_vendor_excel(file_bytes: bytes) -> list[Vendor]:
    """
    Parse vendor list from raw Excel bytes. Extracts ONLY Vendor Name and
    Vendor Website Link columns, regardless of what else is present in the
    sheet. Raises IngestionError with a user-facing message on failure.
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
        ws = wb.active
    except Exception as exc:
        raise IngestionError(f"Could not read the Excel file: {exc}") from exc

    if ws.max_row < 2:
        raise IngestionError("The uploaded Excel file has no rows.")

    header_row_idx = _find_header_row(ws)
    if header_row_idx is None:
        # Fall back to row 1 so the error message below can report what was found
        header_row_idx = 1

    header_values = [cell.value for cell in ws[header_row_idx]]
    name_idx = _find_column_index(header_values, NAME_ALIASES)
    url_idx = _find_column_index(header_values, URL_ALIASES)

    if name_idx is None or url_idx is None:
        found_cols = [v for v in header_values if v is not None]
        raise IngestionError(
            "Could not find required columns. The file must contain a "
            "'Vendor Name' column and a 'Vendor Website Link' column "
            f"(found columns: {found_cols})."
        )

    vendors: list[Vendor] = []
    seen_domains: set[str] = set()

    for row in ws.iter_rows(min_row=header_row_idx + 1):
        if name_idx >= len(row) or url_idx >= len(row):
            continue
        name_cell = row[name_idx]
        url_cell = row[url_idx]

        raw_name = name_cell.value
        if raw_name is None:
            continue
        name = re.sub(r"\s+", " ", str(raw_name).strip())
        if not name:
            continue

        raw_url = _extract_cell_url(url_cell)
        clean_url = _clean_url(raw_url)
        if clean_url is None:
            continue

        domain = _extract_domain(clean_url)
        if domain in seen_domains:
            continue
        seen_domains.add(domain)

        vendors.append(Vendor(name=name, website=clean_url, domain=domain))

    if not vendors:
        raise IngestionError(
            "No valid vendor rows found after cleaning. Check that the "
            "Vendor Website Link column contains valid URLs/domains."
        )

    return vendors
