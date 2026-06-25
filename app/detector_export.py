"""
Detector Results Excel Export
----------------------------------
Builds an .xlsx workbook from a completed ad-hoc detection job's results,
matching the output columns specified in the original feature request:
  - Vendor Name / Domain
  - Selected Detector applied
  - Risk Score / Cyber Security Rating (letter grade)
  - Incident Summary / Description
  - Monitoring Status (Active Continuous vs. Idle Ad-hoc)
  - Timestamp of Last Detection / Alert Delta
"""
from __future__ import annotations

import io
import time

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")

COLUMNS = [
    ("Vendor Name", 22),
    ("Domain", 22),
    ("Detector Applied", 30),
    ("Risk Score", 12),
    ("Rating", 10),
    ("Incident Summary / Description", 60),
    ("Monitoring Status", 18),
    ("Timestamp", 22),
]


def build_export_workbook(detect_job_results: list[dict], monitoring_lookup: dict[str, dict] | None = None) -> bytes:
    """
    detect_job_results: the "results" list from a completed /api/detect job,
        i.e. [{"vendor_name", "domain", "results": [{"detector", ...}, ...]}, ...]
    monitoring_lookup: optional {domain: monitoring_config_dict} to populate
        the Monitoring Status column accurately; if omitted, all rows show
        "Idle (Ad-hoc)".
    """
    monitoring_lookup = monitoring_lookup or {}
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Threat Detection"

    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    row_idx = 2
    for vendor_entry in detect_job_results:
        vendor_name = vendor_entry.get("vendor_name", "")
        domain = vendor_entry.get("domain", "")
        monitoring_config = monitoring_lookup.get(domain)
        if monitoring_config and monitoring_config.get("mode") == "continuous":
            monitoring_status = f"Active Continuous ({monitoring_config.get('frequency', 'daily')})"
        else:
            monitoring_status = "Idle (Ad-hoc)"

        for detector_result in vendor_entry.get("results", []):
            score = detector_result.get("risk_score")
            rating = detector_result.get("rating_letter") or "—"
            summary = detector_result.get("summary", "")
            if detector_result.get("error"):
                summary = f"{summary} [Error: {detector_result['error']}]" if summary else f"Error: {detector_result['error']}"

            values = [
                vendor_name,
                domain,
                detector_result.get("detector_label", detector_result.get("detector", "")),
                score if score is not None else "—",
                rating,
                summary,
                monitoring_status,
                time.strftime("%Y-%m-%d %H:%M:%S UTC", time.gmtime()),
            ]
            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if col_idx == 6:  # summary column gets wrap
                    cell.alignment = WRAP
            row_idx += 1

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
